"""Executive Excel export pack builders for Overview."""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
import re
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.calculations.refill_growth import (
    compute_refill_growth_components_anchor_safe,
    shifted_portfolio_refill_weights,
    t0_portfolio_weights,
)
from src.dashboard.plots.rate_scenario_plots import build_scenario_matrix_table


def default_export_filename(product: str, t1: pd.Timestamp, t2: pd.Timestamp) -> str:
    """Return a deterministic export filename."""
    safe_product = re.sub(r'[^A-Za-z0-9_-]+', '_', str(product or 'product')).strip('_') or 'product'
    t1s = (pd.Timestamp(t1) + pd.offsets.MonthEnd(0)).date().isoformat()
    t2s = (pd.Timestamp(t2) + pd.offsets.MonthEnd(0)).date().isoformat()
    return f'nii_executive_pack_{safe_product}_T1_{t1s}_T2_{t2s}.xlsx'


def _parse_ids(json_text: str) -> list[str]:
    try:
        ids = json.loads(str(json_text or '[]'))
        if isinstance(ids, list):
            return [str(x) for x in ids]
    except Exception:
        pass
    return []


def _curve_rate_by_tenor(
    *,
    tenor_points: pd.Series,
    base_notional_total: pd.Series,
    base_effective_total: pd.Series,
    curve_df: pd.DataFrame | None = None,
    basis_date: pd.Timestamp | None = None,
) -> pd.Series:
    """Return annualized curve rate by tenor with observed-ratio fallback."""
    idx = base_notional_total.index
    eff_rate = pd.Series(0.0, index=idx, dtype=float)
    tenor = pd.Series(tenor_points, index=idx).astype(float).round().astype(int)
    curve_rate = None

    if curve_df is not None and basis_date is not None and not curve_df.empty:
        c = curve_df.copy()
        c.columns = [str(col).strip().lower() for col in c.columns]
        needed_cols = {'ir_date', 'ir_tenor', 'rate'}
        if needed_cols.issubset(c.columns):
            c['ir_date'] = pd.to_datetime(c['ir_date'])
            c['ir_tenor'] = pd.to_numeric(c['ir_tenor'], errors='coerce')
            c['rate'] = pd.to_numeric(c['rate'], errors='coerce')
            c = c.dropna(subset=['ir_date', 'ir_tenor', 'rate']).copy()
            if not c.empty:
                as_of = pd.Timestamp(basis_date) + pd.offsets.MonthEnd(0)
                prior = c[c['ir_date'] <= as_of]
                curve_date = c['ir_date'].min() if prior.empty else prior['ir_date'].max()
                cs = c[c['ir_date'] == curve_date].copy()
                cs = cs.sort_values('ir_tenor').drop_duplicates('ir_tenor', keep='last')
                if len(cs) >= 2:
                    x = cs['ir_tenor'].astype(float).to_numpy()
                    y = cs['rate'].astype(float).to_numpy()
                    q = tenor.astype(float).to_numpy()
                    curve_rate = pd.Series(np.interp(q, x, y), index=idx, dtype=float)
                elif len(cs) == 1:
                    curve_rate = pd.Series(float(cs['rate'].iloc[0]), index=idx, dtype=float)

    if curve_rate is None:
        non_zero = base_notional_total.abs() > 1e-12
        eff_rate.loc[non_zero] = (
            base_effective_total.loc[non_zero].abs() / base_notional_total.loc[non_zero].abs()
        ).astype(float)
    else:
        eff_rate = curve_rate.astype(float)
    return eff_rate


def _monthly_runoff_y1(
    *,
    calendar_runoff: pd.DataFrame,
    t2: pd.Timestamp,
) -> pd.DataFrame:
    t2_me = pd.Timestamp(t2) + pd.offsets.MonthEnd(0)
    y1_months = pd.DatetimeIndex([t2_me + pd.offsets.MonthEnd(k) for k in range(12)])
    if calendar_runoff is None or calendar_runoff.empty:
        return pd.DataFrame({'calendar_month_end': y1_months})

    work = calendar_runoff.copy()
    if 'calendar_month_end' not in work.columns:
        return pd.DataFrame({'calendar_month_end': y1_months})
    work['calendar_month_end'] = pd.to_datetime(work['calendar_month_end'], errors='coerce') + pd.offsets.MonthEnd(0)
    work = work.dropna(subset=['calendar_month_end']).set_index('calendar_month_end').sort_index()
    out = work.reindex(y1_months).fillna(0.0).reset_index().rename(columns={'index': 'calendar_month_end'})
    return out


def _build_distribution_outputs(
    *,
    y1_calendar_runoff: pd.DataFrame,
    runoff_delta: pd.DataFrame | None,
    curve_df: pd.DataFrame | None,
    t2: pd.Timestamp,
    growth_mode: str,
    growth_monthly_value: float,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    notes: list[str] = []
    df = y1_calendar_runoff.copy()
    if df.empty:
        notes.append('No runoff points available for first-year distribution.')
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), notes

    cumulative_col = 'cumulative_signed_notional_t2' if 'cumulative_signed_notional_t2' in df.columns else 'cumulative_abs_notional_t2'
    if cumulative_col not in df.columns:
        notes.append('Missing cumulative notional columns for T2; distribution set to zero.')
        df[cumulative_col] = 0.0

    growth_components = compute_refill_growth_components_anchor_safe(
        cumulative_notional=df[cumulative_col].astype(float),
        growth_mode=growth_mode,
        monthly_growth_amount=float(growth_monthly_value),
    )
    refill_required = growth_components['refill_required'].astype(float).reset_index(drop=True)
    growth_required = growth_components['growth_required'].astype(float).reset_index(drop=True)
    month_ends = pd.to_datetime(df['calendar_month_end'], errors='coerce')

    shifted = shifted_portfolio_refill_weights(runoff_delta)
    t0 = t0_portfolio_weights(runoff_delta, basis='T2')
    refill_w = None
    if shifted is not None and not shifted.empty:
        refill_w = shifted.set_index('tenor')['weight'].astype(float)
    elif t0 is not None and not t0.empty:
        refill_w = t0.astype(float)
    if refill_w is None or float(refill_w.sum()) <= 1e-12:
        notes.append('Could not derive refill tenor weights from runoff comparison data.')
        refill_w = pd.Series(dtype=float)
    else:
        refill_w = refill_w.clip(lower=0.0)
        refill_w = refill_w / float(refill_w.sum())

    include_growth = str(growth_mode or '').strip().lower() == 'user_defined' and float(growth_required.abs().sum()) > 1e-12
    growth_w = pd.Series(dtype=float)
    if include_growth:
        if t0 is not None and not t0.empty and float(t0.sum()) > 1e-12:
            growth_w = t0.astype(float).clip(lower=0.0)
            growth_w = growth_w / float(growth_w.sum())
        elif not refill_w.empty:
            growth_w = refill_w.copy()
        else:
            notes.append('Growth is enabled, but no tenor weights were available for growth allocation.')

    tenor_set: set[int] = set()
    tenor_set.update([int(x) for x in refill_w.index.tolist()]) if not refill_w.empty else None
    tenor_set.update([int(x) for x in growth_w.index.tolist()]) if not growth_w.empty else None
    if not tenor_set:
        month_labels = month_ends.dt.strftime('%Y-%m')
        grid = pd.DataFrame({'Refill Tenor Bucket (Months)': []})
        for label in month_labels.tolist():
            grid[label] = []
    else:
        tenor_index = pd.Index(sorted(tenor_set), dtype=int)
        refill_vec = refill_w.reindex(tenor_index, fill_value=0.0).to_numpy(dtype=float) if not refill_w.empty else np.zeros(len(tenor_index), dtype=float)
        growth_vec = growth_w.reindex(tenor_index, fill_value=0.0).to_numpy(dtype=float) if not growth_w.empty else np.zeros(len(tenor_index), dtype=float)
        z = np.outer(refill_vec, refill_required.abs().to_numpy(dtype=float))
        if include_growth and growth_vec.size:
            z = z + np.outer(growth_vec, growth_required.abs().to_numpy(dtype=float))
        month_labels = month_ends.dt.strftime('%Y-%m')
        grid = pd.DataFrame(z, columns=month_labels.tolist())
        grid.insert(0, 'Refill Tenor Bucket (Months)', tenor_index.to_numpy())

    abs_notional_col = 'abs_notional_t2' if 'abs_notional_t2' in df.columns else 'signed_notional_t2'
    eff_col = 'effective_interest_t2' if 'effective_interest_t2' in df.columns else 'notional_coupon_t2'
    if abs_notional_col not in df.columns:
        df[abs_notional_col] = 0.0
    if eff_col not in df.columns:
        df[eff_col] = 0.0

    tenor_points = pd.Series(np.arange(1, len(df) + 1), dtype=int)
    refill_rate = _curve_rate_by_tenor(
        tenor_points=tenor_points,
        base_notional_total=df[abs_notional_col].astype(float),
        base_effective_total=df[eff_col].astype(float),
        curve_df=curve_df,
        basis_date=pd.Timestamp(t2),
    ).reset_index(drop=True)
    annual_refill_interest = refill_required * refill_rate
    annual_growth_interest = growth_required * refill_rate
    total_volume = refill_required + growth_required
    annual_total_interest = annual_refill_interest + annual_growth_interest
    implied_weighted_rate = pd.Series(np.nan, index=total_volume.index, dtype=float)
    nz = total_volume.abs() > 1e-12
    implied_weighted_rate.loc[nz] = annual_total_interest.loc[nz] / total_volume.loc[nz]

    monthly_summary = pd.DataFrame(
        {
            'Calendar Month End': month_ends,
            'Refill Volume': refill_required,
            'Growth Volume': growth_required,
            'Total Volume': total_volume,
            'Refill Interest (Annualized)': annual_refill_interest,
            'Growth Interest (Annualized)': annual_growth_interest,
            'Total Interest (Annualized)': annual_total_interest,
            'Implied Weighted Rate': implied_weighted_rate,
        }
    )

    legend = pd.DataFrame(
        [
            {'Field': 'Growth Mode', 'Value': str(growth_mode)},
            {'Field': 'Includes Growth', 'Value': bool(include_growth)},
            {'Field': 'Allocation Basis', 'Value': 'Shifted refill weights + T0 growth weights (fallback: shifted/T0)'},
            {'Field': 'Anchor Month', 'Value': (pd.Timestamp(t2) + pd.offsets.MonthEnd(0)).date().isoformat()},
        ]
    )
    return legend, grid, monthly_summary, notes


def build_export_context(
    *,
    path: str,
    product: str,
    t1: pd.Timestamp,
    t2: pd.Timestamp,
    growth_mode: str,
    growth_monthly_value: float,
    scenario_payload_json: str,
    active_ids_json: str,
    overview_metrics: pd.DataFrame | None = None,
    overview_delta_kpis: dict[str, float] | None = None,
    yearly_summary: pd.DataFrame | None = None,
    monthly_base: pd.DataFrame | None = None,
    monthly_scenarios: pd.DataFrame | None = None,
    calendar_runoff: pd.DataFrame | None = None,
    runoff_delta: pd.DataFrame | None = None,
    curve_df: pd.DataFrame | None = None,
) -> dict[str, pd.DataFrame | dict[str, Any]]:
    """Build normalized dataframes for executive export sheets."""
    t1_me = pd.Timestamp(t1) + pd.offsets.MonthEnd(0)
    t2_me = pd.Timestamp(t2) + pd.offsets.MonthEnd(0)
    active_ids = _parse_ids(active_ids_json)
    scenario_records = []
    try:
        parsed = json.loads(str(scenario_payload_json or '[]'))
        if isinstance(parsed, list):
            scenario_records = parsed
    except Exception:
        scenario_records = []

    notes: list[str] = []

    if overview_metrics is None:
        overview_metrics_df = pd.DataFrame(columns=['Category', 'Metric', 'T1', 'T2', 'Delta'])
    else:
        tmp = overview_metrics.copy().reset_index()
        metric_col = None
        for candidate in ['Metric', 'metric', 'index']:
            if candidate in tmp.columns:
                metric_col = candidate
                break
        if metric_col is None:
            metric_col = str(tmp.columns[0]) if len(tmp.columns) > 0 else 'Metric'
        if metric_col != 'Metric':
            tmp = tmp.rename(columns={metric_col: 'Metric'})
        if 'Metric' not in tmp.columns:
            tmp['Metric'] = ''
        col_t1 = next((c for c in tmp.columns if str(c).startswith('T1 ')), 'T1')
        col_t2 = next((c for c in tmp.columns if str(c).startswith('T2 ')), 'T2')
        col_d = next((c for c in tmp.columns if str(c).startswith('Delta')), 'Delta')
        for missing in [col_t1, col_t2, col_d]:
            if missing not in tmp.columns:
                tmp[missing] = np.nan
        overview_metrics_df = pd.DataFrame(
            {
                'Category': 'Metrics Table',
                'Metric': tmp['Metric'].astype(str),
                'T1': pd.to_numeric(tmp[col_t1], errors='coerce'),
                'T2': pd.to_numeric(tmp[col_t2], errors='coerce'),
                'Delta': pd.to_numeric(tmp[col_d], errors='coerce'),
            }
        )

    kpi_rows = []
    for k, v in (overview_delta_kpis or {}).items():
        kpi_rows.append({'Category': 'Delta KPI', 'Metric': str(k), 'T1': np.nan, 'T2': np.nan, 'Delta': float(v)})
    if kpi_rows:
        overview_metrics_df = pd.concat([pd.DataFrame(kpi_rows), overview_metrics_df], ignore_index=True)

    yearly_summary = yearly_summary if yearly_summary is not None else pd.DataFrame()
    monthly_base = monthly_base if monthly_base is not None else pd.DataFrame()
    monthly_scenarios = monthly_scenarios if monthly_scenarios is not None else pd.DataFrame()
    scenario_delta = build_scenario_matrix_table(yearly_summary, view_mode='delta')
    scenario_absolute = build_scenario_matrix_table(
        yearly_summary,
        view_mode='absolute',
        monthly_base=monthly_base,
        monthly_scenarios=monthly_scenarios,
    )
    if scenario_delta.empty:
        notes.append('Scenario delta matrix unavailable for current selection.')
    if scenario_absolute.empty:
        notes.append('Scenario absolute matrix unavailable for current selection.')

    y1_runoff = _monthly_runoff_y1(
        calendar_runoff=(calendar_runoff if calendar_runoff is not None else pd.DataFrame()),
        t2=t2_me,
    )
    dist_legend, dist_grid, dist_summary, dist_notes = _build_distribution_outputs(
        y1_calendar_runoff=y1_runoff,
        runoff_delta=runoff_delta,
        curve_df=curve_df,
        t2=t2_me,
        growth_mode=growth_mode,
        growth_monthly_value=growth_monthly_value,
    )
    notes.extend(dist_notes)

    metadata = pd.DataFrame(
        [
            {'Field': 'Generated At', 'Value': datetime.now().isoformat(timespec='seconds')},
            {'Field': 'Workbook Path', 'Value': str(path)},
            {'Field': 'Product', 'Value': str(product)},
            {'Field': 'T1', 'Value': t1_me.date().isoformat()},
            {'Field': 'T2', 'Value': t2_me.date().isoformat()},
            {'Field': 'Growth Mode', 'Value': str(growth_mode)},
            {'Field': 'Growth Monthly Value', 'Value': float(growth_monthly_value)},
            {'Field': 'Scenario Set Count', 'Value': int(len(scenario_records))},
            {'Field': 'Active Scenario IDs', 'Value': ', '.join(active_ids)},
            {'Field': 'Report Scope', 'Value': 'Executive Pack'},
            {'Field': 'Report Version', 'Value': '1'},
            {'Field': 'Notes', 'Value': ' | '.join(notes) if notes else ''},
        ]
    )

    return {
        'summary_metadata': metadata,
        'overview_metrics': overview_metrics_df,
        'scenario_matrix_delta': scenario_delta,
        'scenario_matrix_absolute': scenario_absolute,
        'distribution_legend': dist_legend,
        'distribution_month_tenor': dist_grid,
        'distribution_monthly_summary': dist_summary,
    }


def _format_worksheet(
    ws,
    *,
    header_row: int = 1,
    freeze_panes: str = 'A2',
) -> None:
    ws.freeze_panes = freeze_panes
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 0 or max_col <= 0:
        return

    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)

    headers: dict[int, str] = {}
    for col_idx in range(1, max_col + 1):
        headers[col_idx] = str(ws.cell(row=header_row, column=col_idx).value or '').strip().lower()

    for row_idx in range(header_row + 1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (pd.Timestamp, datetime)):
                cell.number_format = 'YYYY-MM-DD'
                continue
            if isinstance(cell.value, bool):
                continue
            if isinstance(cell.value, (int, float, np.integer, np.floating)):
                header = headers.get(col_idx, '')
                if 'rate' in header or '%' in header:
                    cell.number_format = '0.0000%'
                elif 'count' in header:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, min(max_row, 200) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            text = '' if val is None else str(val)
            max_len = max(max_len, len(text))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)


def build_export_workbook_bytes(context: dict[str, Any], *, workbook_title: str) -> bytes:
    """Serialize export context into an Excel workbook."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(context.get('summary_metadata', pd.DataFrame())).to_excel(
            writer, sheet_name='Summary_Metadata', index=False
        )
        pd.DataFrame(context.get('overview_metrics', pd.DataFrame())).to_excel(
            writer, sheet_name='Overview_Metrics', index=False
        )
        pd.DataFrame(context.get('scenario_matrix_delta', pd.DataFrame())).to_excel(
            writer, sheet_name='Scenario_Matrix_Delta', index=False
        )
        pd.DataFrame(context.get('scenario_matrix_absolute', pd.DataFrame())).to_excel(
            writer, sheet_name='Scenario_Matrix_Absolute', index=False
        )

        dist_legend = pd.DataFrame(context.get('distribution_legend', pd.DataFrame()))
        dist_grid = pd.DataFrame(context.get('distribution_month_tenor', pd.DataFrame()))
        dist_start_row = len(dist_legend) + 2
        if not dist_legend.empty:
            dist_legend.to_excel(
                writer,
                sheet_name='Distribution_Month_Tenor_Y1',
                index=False,
                startrow=0,
            )
        dist_grid.to_excel(
            writer,
            sheet_name='Distribution_Month_Tenor_Y1',
            index=False,
            startrow=dist_start_row,
        )
        pd.DataFrame(context.get('distribution_monthly_summary', pd.DataFrame())).to_excel(
            writer, sheet_name='Distribution_Monthly_Summary_Y1', index=False
        )

        wb = writer.book
        wb.properties.title = str(workbook_title)

        _format_worksheet(writer.sheets['Summary_Metadata'])
        _format_worksheet(writer.sheets['Overview_Metrics'])
        _format_worksheet(writer.sheets['Scenario_Matrix_Delta'])
        _format_worksheet(writer.sheets['Scenario_Matrix_Absolute'])
        header_row = dist_start_row + 1
        _format_worksheet(
            writer.sheets['Distribution_Month_Tenor_Y1'],
            header_row=header_row,
            freeze_panes=f'A{header_row + 1}',
        )
        _format_worksheet(writer.sheets['Distribution_Monthly_Summary_Y1'])

    return output.getvalue()
